import io
import json
import os
import subprocess
import sys
import threading
import time
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from flask import Flask, Response, jsonify, redirect, render_template, request, send_file, session, stream_with_context, url_for

sys.path.insert(0, str(Path(__file__).parent))

import config
from config import SEARCH_RESULT_CAP, SECRET_KEY
from connectors.synced_file import SyncedFileSource
from core import logger as log
from core import snapshot_store, source_status
from core.view_groups import group_views_by_source, ViewGroup
from core.fileio import friendly_read_error, is_csv, read_shared_bytes
from core.join import left_join as join_left_merge
from core.normalize import normalize_key, parse_values
from core.poller import start_poller
from core.search import search
from core.settings_store import load_settings, save_settings
from core.templates_store import (
    delete_template,
    list_templates,
    load_template,
    save_template,
    validate_template,
)
from version import RELEASE_VERSION, SERVER_VERSION, UI_VERSION

app = Flask(__name__)
app.secret_key = SECRET_KEY
app.config["TEMPLATES_AUTO_RELOAD"] = True  # pick up template edits without a restart


@app.context_processor
def inject_versions():
    return {"server_version": SERVER_VERSION, "ui_version": UI_VERSION,
            "release_version": RELEASE_VERSION}


@app.context_processor
def inject_settings():
    return {"settings": load_settings()}

# In-memory source cache: path -> SyncedFileSource
_sources: dict[str, SyncedFileSource] = {}


def _header_row_to_pandas(header_row_1based: int) -> int:
    """Convert 1-based Excel row number to 0-based pandas header index."""
    return max(0, int(header_row_1based) - 1)


def _get_source(
    template: dict,
    sheet_name: str | None = None,
    table_name: str | None = None,
    _override: bool = False,
) -> SyncedFileSource:
    """Get (or lazily build) a source for a template.

    When _override is True, uses the passed sheet_name/table_name instead
    of the template's primary source values — used to resolve per-view
    overrides (schema v4). The cache key includes sheet/table so different
    sheets of the same workbook are cached as distinct source instances.
    """
    src = template["source"]
    stype = (src.get("type") or "local").strip()
    header_row = _header_row_to_pandas(src.get("header_row", 1))
    if _override:
        use_sheet = sheet_name
        use_table = table_name
    else:
        use_sheet = src.get("sheet_name") or None
        use_table = src.get("table_name") or None

    if stype == "sharepoint":
        from connectors.sharepoint_cached import SharePointCachedSource
        key = f"sp::{src.get('drive_id')}::{src.get('item_id')}::{use_sheet or ''}::{use_table or ''}"
        if key not in _sources:
            _sources[key] = SharePointCachedSource(
                drive_id=src["drive_id"],
                item_id=src["item_id"],
                name=src.get("name") or "file.xlsx",
                sheet_name=use_sheet,
                table_name=use_table,
                header_row=header_row,
                template_name=template.get("name", ""),
            )
        return _sources[key]

    if stype == "sql":
        from connectors.sql_query import SqlQuerySource
        from core import sql_connections, sql_credentials
        conn_id = src.get("connection_id") or ""
        query = src.get("query") or ""
        key = f"sql::{conn_id}::{hash(query)}"
        if key not in _sources:
            conn = sql_connections.load_connection(conn_id)
            if conn is None:
                raise ValueError(
                    f"SQL connection '{conn_id}' not found. "
                    f"Recreate it in the template builder."
                )
            cred = sql_credentials.load_credential(conn.get("credential_id", ""))
            if cred is None:
                raise ValueError(
                    "SQL credential missing or unreadable "
                    "(re-save this connection with your password)."
                )
            username, password = cred
            _sources[key] = SqlQuerySource(
                connection={**conn, "username": username, "password": password},
                query=query,
            )
        return _sources[key]

    path = src["path"]
    key = f"{path}::{use_sheet or ''}::{use_table or ''}"
    if key not in _sources:
        _sources[key] = SyncedFileSource(
            path=path,
            sheet_name=use_sheet,
            table_name=use_table,
            header_row=header_row,
        )
    return _sources[key]


def _get_view_group_source(template: dict, group: ViewGroup) -> SyncedFileSource:
    """Resolve the DataSource for a specific view group."""
    if group.is_primary:
        return _get_source(template)
    return _get_source(
        template,
        sheet_name=group.sheet_name,
        table_name=group.table_name,
        _override=True,
    )


def _ensure_loaded(source: SyncedFileSource) -> None:
    if source.dataframe is None:
        source.load()


def _format_col_list(cols: list[str]) -> str:
    if not cols:
        return ""
    if len(cols) == 1:
        return cols[0]
    if len(cols) == 2:
        return f"{cols[0]} and {cols[1]}"
    return ", ".join(cols[:-1]) + f", and {cols[-1]}"


def _key_col_header(column_queries: list, labels: dict) -> str:
    """Header for the match column: the searched field's name when exactly one
    field was used, otherwise 'Match' (multiple fields → composite match string)."""
    if len(column_queries) == 1:
        col = column_queries[0][0]
        return labels.get(col, col)
    return "Match"


def _search_summary(template: dict | None) -> str:
    if not template:
        return ""
    key_cols = template.get("key_columns", [])
    result_cols = template.get("result_columns", [])
    if not key_cols:
        return ""
    key_str = _format_col_list(key_cols)
    result_str = _format_col_list(result_cols) if result_cols else "all columns"
    return f"Search by {key_str}; returns {result_str}."


def _apply_default_filter(df, template: dict):
    df_filter = template.get("default_filter")
    if not df_filter or not df_filter.get("column"):
        return df
    col = df_filter["column"]
    val = df_filter.get("equals", "")
    if col not in df.columns:
        return df
    return df[df[col].apply(normalize_key) == normalize_key(str(val))]


# ---------------------------------------------------------------------------
# Search screen
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    templates = list_templates()
    selected_name = request.args.get("template", "")
    selected = None
    warnings = []
    if selected_name:
        try:
            selected = load_template(selected_name)
            # Only validate against live columns if the primary source is already cached.
            # Do NOT load the workbook here — that happens on first search.
            try:
                primary = _get_source(selected)
                if primary.dataframe is not None:
                    warnings = validate_template(selected, primary.columns())
                else:
                    warnings = validate_template(selected)  # schema-only, no column check
            except Exception:
                warnings = validate_template(selected)
        except Exception as e:
            warnings = [str(e)]
    auto_check = session.get("auto_check", False)

    form_key_values: list[str] = []
    if selected:
        for i in range(len(selected.get("key_columns", []))):
            form_key_values.append(request.args.get(f"key_{i}", ""))
    form_mode = request.args.get("mode", "exact")
    auto_run = request.args.get("run") == "1" and any(v.strip() for v in form_key_values)

    return render_template(
        "search_c.html",
        templates=templates,
        selected=selected,
        selected_name=selected_name,
        warnings=warnings,
        auto_check=auto_check,
        form_key_values=form_key_values,
        form_mode=form_mode,
        search_summary=_search_summary(selected),
        auto_run=auto_run,
    )


@app.route("/api/search", methods=["POST"])
def api_search():
    """SSE stream: emits status events during search, then the final JSON result."""
    template_name = request.form.get("template", "")
    mode = request.form.get("mode", "exact")
    force_reload = request.form.get("force_reload") == "1"
    auto_check = session.get("auto_check", False)

    key_raw: list[str] = []
    for i in range(20):
        v = request.form.get(f"key_{i}")
        if v is None:
            break
        key_raw.append(v)

    # The client sends back the snapshot ids its previous search created (it's
    # the only thing that still holds them — see the snapshot_ids comment
    # below) so we can delete them instead of leaving them on disk until the
    # next TTL sweep.
    try:
        prev_snapshot_ids = json.loads(request.form.get("prev_snapshot_ids", "[]"))
    except (TypeError, ValueError):
        prev_snapshot_ids = []
    for sid in prev_snapshot_ids:
        if sid:
            snapshot_store.delete_snapshot(sid)

    def generate():
        def _evt(event, data):
            return f"event: {event}\ndata: {json.dumps(data)}\n\n"

        try:
            template = load_template(template_name)
            if not template:
                yield _evt("error", {"message": f"Template '{template_name}' not found."})
                return

            key_columns = template["key_columns"]
            labels = template.get("labels", {})
            column_queries: list[tuple[str, list[str]]] = []
            for i, col in enumerate(key_columns):
                raw = key_raw[i] if i < len(key_raw) else ""
                vals = parse_values(raw)
                if vals:
                    column_queries.append((col, vals))

            groups = group_views_by_source(template)
            src_name = template.get("source", {}).get("name") or template.get("source", {}).get("path") or template_name

            new_snapshot_ids: dict[str, str] = {}
            group_results: list[dict] = []
            primary_source = None
            primary_warnings: list[str] = []

            t0 = time.monotonic()
            for gi, group in enumerate(groups):
                if group.join_sheet_name:
                    # Merged view: base sheet LEFT JOIN a second sheet of the
                    # same workbook. Load both, merge, then search/display the
                    # combined frame exactly like a normal group.
                    base_source = _get_source(
                        template, sheet_name=group.sheet_name,
                        table_name=group.table_name, _override=True,
                    )
                    join_source = _get_source(
                        template, sheet_name=group.join_sheet_name, _override=True,
                    )
                    source = base_source
                    needs_load = (
                        force_reload
                        or (auto_check and (base_source.is_stale() or join_source.is_stale()))
                        or base_source.dataframe is None or join_source.dataframe is None
                    )
                    if needs_load:
                        yield _evt("status", {"stage": "loading", "detail": f"Loading {src_name}…"})
                        base_source.load()
                        join_source.load()
                    df = join_left_merge(
                        base_source.dataframe, join_source.dataframe,
                        group.join_on, group.join_sheet_name,
                    )
                    yield _evt("status", {"stage": "searching", "detail": f"Searching {len(df):,} rows…"})
                    df = _apply_default_filter(df, template)
                else:
                    source = _get_view_group_source(template, group)
                    if group.is_primary:
                        primary_source = source

                    needs_load = force_reload or (auto_check and source.is_stale()) or source.dataframe is None
                    if needs_load:
                        yield _evt("status", {"stage": "loading", "detail": f"Loading {src_name}…"})
                        source.load()

                    yield _evt("status", {"stage": "searching", "detail": f"Searching {len(source.dataframe):,} rows…"})

                    df = source.dataframe.copy()
                    df = _apply_default_filter(df, template)

                # Only flag a sheet as unsearchable for columns actually being
                # queried this search — a sheet lacking one configured key
                # (e.g. PropertyID) is still fully searchable by another
                # (e.g. Address) as long as that column exists here and the
                # user filled it in.
                queried_cols = [c for c, _ in column_queries]
                missing_keys = [c for c in queried_cols if c not in df.columns]
                if missing_keys:
                    group_results.append({
                        "group_key": group.key, "sheet_name": group.sheet_name,
                        "table_name": group.table_name, "is_primary": group.is_primary,
                        "views": group.views,
                        "disabled_reason": f"sheet '{group.sheet_name or '(default)'}' has no " + ", ".join(f"'{c}'" for c in missing_keys) + " column",
                        "display_rows": [], "all_view_cols": [], "total_matches": 0,
                        "not_found": [], "truncated": False,
                        "source_timestamp": source.source_timestamp().strftime("%Y-%m-%d %H:%M:%S"),
                    })
                    continue

                sr = search(df, column_queries, mode, limit=SEARCH_RESULT_CAP)

                group_all_cols: list[str] = []
                seen: set[str] = set()
                for v in group.views:
                    for c in v["columns"]:
                        if c not in seen:
                            group_all_cols.append(c)
                            seen.add(c)

                display_rows = sr.rows.copy()
                for c in group_all_cols:
                    if c not in display_rows.columns:
                        display_rows[c] = ""
                display_rows = display_rows[group_all_cols + ["_matched_on", "_duplicate", "_card_title"]]

                sid = snapshot_store.save_snapshot(
                    sr.full_rows, template_name=f"{template_name}__{group.key}",
                    result_columns=group_all_cols, not_found=sr.not_found,
                )
                new_snapshot_ids[group.key] = sid

                rows_json = []
                for _, row in display_rows.iterrows():
                    r = {c: row[c] for c in group_all_cols}
                    r["_matched_on"] = row["_matched_on"]
                    r["_duplicate"] = bool(row["_duplicate"])
                    r["_card_title"] = row["_card_title"]
                    rows_json.append(r)

                group_results.append({
                    "group_key": group.key, "sheet_name": group.sheet_name,
                    "table_name": group.table_name, "is_primary": group.is_primary,
                    "views": group.views, "disabled_reason": None,
                    "display_rows": rows_json, "all_view_cols": group_all_cols,
                    "total_matches": sr.total_matches, "not_found": sr.not_found,
                    "truncated": sr.truncated,
                    "source_timestamp": source.source_timestamp().strftime("%Y-%m-%d %H:%M:%S"),
                })

                if group.is_primary:
                    primary_warnings = validate_template(template, source.columns())

            duration_ms = int((time.monotonic() - t0) * 1000)

            primary_result = next(
                (g for g in group_results if g["is_primary"] and g["disabled_reason"] is None), None,
            )
            log.log_search(
                template_name=template_name, mode=mode,
                value_count=len(column_queries),
                match_count=primary_result["total_matches"] if primary_result else 0,
                not_found=primary_result["not_found"] if primary_result else [],
                duration_ms=duration_ms,
            )

            card_max = load_settings()["card_max"]
            primary_total = primary_result["total_matches"] if primary_result else 0
            view = "cards" if 0 < primary_total <= card_max else "table"

            overall_not_found = primary_result["not_found"] if primary_result else []
            overall_truncated = any(g.get("truncated") for g in group_results)

            primary_group = next((g for g in group_results if g["is_primary"]), group_results[0])

            views_flat = []
            for g in group_results:
                for v in g["views"]:
                    views_flat.append({**v, "group_key": g["group_key"]})

            result_data = {
                "groups": group_results,
                "primary_group_key": primary_group["group_key"],
                "views": views_flat,
                "labels": labels,
                "total_matches": primary_total,
                "not_found": overall_not_found,
                "truncated": overall_truncated,
                "queried_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "source_timestamp": (
                    primary_source.source_timestamp().strftime("%Y-%m-%d %H:%M:%S")
                    if primary_source else datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ),
                "view": view,
                "duration_ms": duration_ms,
            }

            key_col_header = _key_col_header(column_queries, labels)

            yield _evt("status", {"stage": "rendering", "detail": "Rendering results…"})
            yield _evt("result", {
                "result": result_data,
                "warnings": primary_warnings,
                "key_col_header": key_col_header,
                "template_name": template_name,
                "links": template.get("links", []),
                # Session writes inside this generator never reach the client:
                # Flask commits the session cookie header before a streaming
                # response's body runs. Send snapshot ids in-band instead; the
                # client passes them explicitly to /api/more_rows and /export.
                "snapshot_ids": new_snapshot_ids,
            })

        except Exception as e:
            yield _evt("error", {"message": friendly_read_error(e)})

    return Response(stream_with_context(generate()), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/refresh", methods=["POST"])
def do_refresh():
    template_name = request.form.get("template", "")
    try:
        template = load_template(template_name)
        source = _get_source(template)
        source.load()
        log.log_refresh(template["source"]["path"])
    except Exception:
        pass
    return redirect(url_for("index", template=template_name))


@app.route("/export", methods=["POST"])
def do_export():
    # The search response returns snapshot ids in-band (not via the session —
    # Flask commits the session cookie before a streaming response's generator
    # body runs, so a mid-stream session write is silently lost). The client
    # holds them and sends the active group's id explicitly.
    sid = request.form.get("snapshot_id", "")
    snap = snapshot_store.load_snapshot(sid) if sid else None
    if snap is None:
        return redirect(url_for("index"))

    full_df = snap["df"]
    result_columns = snap["result_columns"]
    export_cols = [c for c in result_columns if c in full_df.columns]
    export_df = full_df[export_cols] if export_cols else full_df

    config.EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in snap["template_name"])
    filename = f"{safe_name}_{ts}.csv"
    filepath = config.EXPORTS_DIR / filename

    export_df.to_csv(filepath, index=False, encoding="utf-8-sig")
    nf = snap.get("not_found", [])
    if nf:
        with open(filepath, "a", encoding="utf-8-sig") as f:
            f.write("\n")
            f.write("NOT FOUND\n")
            for v in nf:
                f.write(v + "\n")
    return send_file(filepath, as_attachment=True, download_name=filename)


@app.route("/settings", methods=["POST"])
def update_settings():
    template_name = request.form.get("template", "")
    try:
        payload = {k: request.form.get(k, "") for k in
                   ("card_max", "poll_minutes", "graph_client_id", "graph_tenant")}
        saved = save_settings(payload)
        # Never log the tenant ID or Client ID (may be sensitive)
        log.log_settings_change({k: v for k, v in saved.items() if k != "graph_client_id"})
        return jsonify({"ok": True, "settings": {k: v for k, v in saved.items()
                                                 if k != "graph_client_id"}})
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    return redirect(url_for("index", template=template_name))


@app.route("/api/column_values")
def api_column_values():
    """Distinct values for a key column — used by the search autocomplete.

    Optional query string `q` filters values (case-insensitive substring) so
    large columns don't get truncated to an alphabetic slice before the user
    sees anything matching what they typed.
    """
    template_name = request.args.get("template", "")
    col = request.args.get("col", "")
    q = request.args.get("q", "").strip().lower()
    try:
        template = load_template(template_name)
        if not template or col not in template.get("key_columns", []):
            return jsonify([])
        source = _get_source(template)
        df = source.load()
        if col not in df.columns:
            return jsonify([])
        vals = df[col].dropna().astype(str).unique().tolist()
        if q:
            # Substring, case-insensitive; keep prefix-matches first so the
            # client's ordering feels natural.
            starts, contains = [], []
            for v in vals:
                lv = v.lower()
                if lv.startswith(q):
                    starts.append(v)
                elif q in lv:
                    contains.append(v)
            starts.sort()
            contains.sort()
            return jsonify((starts + contains)[:200])
        # No filter: return the first 2000 alphabetically (unchanged behavior).
        return jsonify(sorted(vals)[:2000])
    except Exception:
        return jsonify([])


@app.route("/api/source_status")
def api_source_status():
    """Lightweight status for the update banner. Never loads a workbook."""
    template_name = request.args.get("template", "")
    out = {"stale": False, "last_checked": None, "last_updated": None,
           "last_error": None, "signed_in": False}
    try:
        template = load_template(template_name)
    except Exception:
        return jsonify(out)
    st = source_status.get_status(template_name)
    out["last_checked"] = st.get("last_checked")
    out["last_updated"] = st.get("last_updated")
    out["last_error"] = st.get("last_error")
    # Auth status for SharePoint sources
    try:
        from core import graph_auth
        out["signed_in"] = graph_auth.auth_state().get("signed_in", False)
    except Exception:
        pass
    src_obj = None
    src_cfg = template.get("source", {}) or {}
    stype = (src_cfg.get("type") or "local").strip()
    if stype == "sharepoint":
        key = f"sp::{src_cfg.get('drive_id')}::{src_cfg.get('item_id')}"
        src_obj = _sources.get(key)
    else:
        src_obj = _sources.get(src_cfg.get("path") or "")
    if src_obj is not None and src_obj.dataframe is not None:
        try:
            out["stale"] = bool(src_obj.is_stale())
        except Exception:
            pass
    return jsonify(out)


# ---------------------------------------------------------------------------
# Microsoft Graph auth (Phase 2) — SharePoint sign-in / status / sign-out
# ---------------------------------------------------------------------------

@app.route("/auth/signin", methods=["POST"])
def auth_signin():
    from core import graph_auth
    if not graph_auth.is_configured():
        return jsonify({"error": "Not configured. Enter your Azure Client ID in Settings."}), 400
    graph_auth.begin_interactive_signin()
    return jsonify({"started": True}), 202


@app.route("/auth/signout", methods=["POST"])
def auth_signout():
    from core import graph_auth
    graph_auth.sign_out()
    return jsonify({"ok": True})


@app.route("/api/auth_status")
def api_auth_status():
    from core import graph_auth
    return jsonify(graph_auth.auth_state())


@app.route("/api/auth_test", methods=["POST"])
def api_auth_test():
    from core import graph_auth, graph_client
    if not graph_auth.is_configured():
        return jsonify({"ok": False, "stage": "config", "message": "No Client ID set."})
    token = graph_auth.get_token_silent()
    if not token:
        return jsonify({"ok": False, "stage": "signin", "message": "Not signed in. Click Sign in first."})
    try:
        name = graph_client.whoami(token)
        return jsonify({"ok": True, "message": f"Connected as {name}."})
    except graph_client.GraphError as e:
        return jsonify({"ok": False, "stage": "graph", "message": e.message})


@app.route("/api/resolve_source", methods=["POST"])
def api_resolve_source():
    """Resolve a SharePoint URL and download it to the cache once."""
    from core import graph_auth, graph_client, source_sync
    payload = request.get_json(silent=True) or {}
    url = (payload.get("url") or "").strip()
    if not url:
        return jsonify({"error": "URL is required."}), 400
    if not graph_auth.is_configured():
        return jsonify({"error": "SharePoint isn't set up yet. Enter your Azure Client ID in Settings.", "needs_config": True}), 400
    token = graph_auth.get_token_silent()
    if not token:
        return jsonify({"error": "Not signed in to Microsoft. Sign in first.", "needs_signin": True}), 401
    try:
        ref = graph_client.resolve_share_url(url, token)
        # First download into the cache
        source_sync.sync_sharepoint_source({
            "type": "sharepoint",
            "drive_id": ref.drive_id, "item_id": ref.item_id, "name": ref.name,
        }, template_name=f"__resolve__{ref.item_id[:8]}")
        cache_path = source_sync.cache_path_for(ref.item_id, ref.name)
        if not cache_path.exists():
            return jsonify({"error": "File could not be downloaded from SharePoint."}), 500
        return jsonify({
            "drive_id": ref.drive_id, "item_id": ref.item_id, "name": ref.name,
            "etag": ref.etag, "cache_path": str(cache_path),
        })
    except graph_client.GraphError as e:
        code = 401 if e.kind == "auth" else 403 if e.kind == "forbidden" else 404 if e.kind == "not_found" else 400
        return jsonify({"error": e.message, "kind": e.kind}), code


@app.route("/api/log_tail")
def api_log_tail():
    from collections import deque
    n = min(int(request.args.get("lines", 200)), 1000)
    try:
        with open(config.LOG_FILE, "r", encoding="utf-8", errors="replace") as f:
            lines = list(deque(f, maxlen=n))
    except (OSError, FileNotFoundError):
        lines = []
    return jsonify({"lines": [l.rstrip("\n") for l in lines]})


@app.route("/api/cross_search", methods=["POST"])
def api_cross_search():
    data = request.get_json(silent=True) or {}
    value = (data.get("value") or "").strip()
    mode = data.get("mode", "exact")
    if not value:
        return jsonify({"error": "Value is required."}), 400
    results = []
    for t in list_templates():
        name = t.get("name", "")
        stype = (t.get("source", {}).get("type") or "local").strip()
        if stype == "sql":
            results.append({"template": name, "skipped": "SQL source", "matches": 0, "column": "", "sample": []})
            continue
        for col in t.get("key_columns", []):
            try:
                source = _get_source(t)
                _ensure_loaded(source)
                df = source.dataframe.copy()
                df = _apply_default_filter(df, t)
                sr = search(df, [(col, [value])], mode, limit=5)
                sample = []
                if sr.total_matches > 0 and "_matched_on" in sr.rows.columns:
                    sample = sr.rows["_matched_on"].head(3).tolist()
                results.append({
                    "template": name, "column": col,
                    "matches": sr.total_matches, "sample": sample,
                })
            except Exception as e:
                results.append({"template": name, "column": col, "matches": 0, "error": str(e), "sample": []})
    return jsonify({"results": results, "value": value, "mode": mode})


@app.route("/api/more_rows")
def api_more_rows():
    """Return one page of a group's full snapshot. Rows are shaped exactly like
    the initial SSE payload's display_rows (flat {col: value, ...} dicts plus
    _matched_on/_duplicate/_card_title) so the client can render a page fetched
    here with the exact same code path as the first page."""
    offset = int(request.args.get("offset", 0))
    limit = int(request.args.get("limit", SEARCH_RESULT_CAP))
    # The search response returns snapshot ids in-band (not via the session —
    # Flask commits the session cookie before a streaming response's generator
    # body runs, so a mid-stream session write is silently lost). The client
    # holds them and sends the active group's id explicitly.
    sid = request.args.get("snapshot_id", "")
    if not sid:
        return jsonify({"error": "Results expired — run the search again."}), 410
    snap = snapshot_store.load_snapshot(sid)
    if snap is None:
        return jsonify({"error": "Results expired — run the search again."}), 410
    df = snap["df"]
    result_columns = snap["result_columns"]
    sl = df.iloc[offset:offset + limit]
    rows = []
    for _, r in sl.iterrows():
        row = {c: (str(r[c]) if c in r.index and pd.notna(r[c]) else "") for c in result_columns}
        row["_matched_on"] = str(r["_matched_on"]) if "_matched_on" in r.index else ""
        row["_duplicate"] = bool(r["_duplicate"]) if "_duplicate" in r.index else False
        row["_card_title"] = str(r["_card_title"]) if "_card_title" in r.index else ""
        rows.append(row)
    return jsonify({
        "rows": rows,
        "columns": result_columns,
        "total": len(df),
        "has_more": offset + limit < len(df),
        "offset": offset,
    })


@app.route("/api/update_check")
def api_update_check():
    from core.update_check import check_update
    return jsonify(check_update(config.UPDATE_REPO, RELEASE_VERSION))


@app.route("/toggle_auto_check", methods=["POST"])
def toggle_auto_check():
    session["auto_check"] = not session.get("auto_check", False)
    template_name = request.form.get("template", "")
    return redirect(url_for("index", template=template_name))


# ---------------------------------------------------------------------------
# Template builder — JSON API (called by client-side JS)
# ---------------------------------------------------------------------------

def _browse_file(title: str, filter_str: str):
    """Open native Windows file picker (forced topmost) and return selected path as JSON."""
    ps_script = (
        "Add-Type -AssemblyName System.Windows.Forms; "
        # Invisible owner window forces the dialog to the front of all other windows
        "$owner = New-Object System.Windows.Forms.Form; "
        "$owner.TopMost = $true; "
        "$owner.WindowState = [System.Windows.Forms.FormWindowState]::Minimized; "
        "$owner.Show(); "
        "$d = New-Object System.Windows.Forms.OpenFileDialog; "
        f"$d.Title = '{title}'; "
        f"$d.Filter = '{filter_str}'; "
        "$d.InitialDirectory = [Environment]::GetFolderPath('UserProfile'); "
        "$result = $d.ShowDialog($owner); "
        "$owner.Dispose(); "
        "if ($result -eq [System.Windows.Forms.DialogResult]::OK) { Write-Output $d.FileName }"
    )
    try:
        result = subprocess.run(
            ["powershell", "-Command", ps_script],
            capture_output=True, text=True, timeout=120,
        )
        path = result.stdout.strip()
        return jsonify({"path": path})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/browse_file")
def api_browse_file():
    return _browse_file(
        title="Select Data File",
        filter_str="Data files (*.xlsx;*.xlsm;*.xls;*.csv)|*.xlsx;*.xlsm;*.xls;*.csv|"
                   "CSV files (*.csv)|*.csv|Excel files (*.xlsx;*.xlsm;*.xls)|*.xlsx;*.xlsm;*.xls|"
                   "All files (*.*)|*.*",
    )


@app.route("/api/sheets", methods=["POST"])
def api_sheets():
    """Return sheet names for a workbook path. CSV files have no sheets."""
    path = request.json.get("path", "")
    if is_csv(path):
        return jsonify({"sheets": [], "is_csv": True})
    try:
        data = read_shared_bytes(path)
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        return jsonify({"sheets": sheets, "is_csv": False})
    except Exception as e:
        return jsonify({"error": friendly_read_error(e)}), 400


@app.route("/api/tables", methods=["POST"])
def api_tables():
    """Return Excel Table names for a given sheet — reads ZIP metadata only, no full workbook load."""
    data = request.json
    path = data.get("path", "")
    sheet_name = data.get("sheet_name") or None
    if is_csv(path):
        return jsonify({"tables": []})
    try:
        file_bytes = read_shared_bytes(path)
        tables = _tables_for_sheet_from_zip(file_bytes, sheet_name)
        return jsonify({"tables": tables})
    except Exception as e:
        return jsonify({"error": friendly_read_error(e)}), 400


def _tables_for_sheet_from_zip(file_bytes: bytes, sheet_name: str | None) -> list[str]:
    """List named Excel Tables for a sheet by reading the xlsx ZIP directly."""
    tables: list[str] = []
    with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
        names = zf.namelist()

        # If sheet_name given, find which worksheet file corresponds to it
        target_sheet_file: str | None = None
        if sheet_name:
            wb_sheet_rids: dict[str, str] = {}
            if "xl/workbook.xml" in names:
                with zf.open("xl/workbook.xml") as f:
                    root = ET.parse(f).getroot()
                    for sh in root.findall(".//{*}sheet"):
                        rid = sh.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", "")
                        wb_sheet_rids[sh.get("name", "")] = rid
            rid_to_file: dict[str, str] = {}
            wb_rels = "xl/_rels/workbook.xml.rels"
            if wb_rels in names:
                with zf.open(wb_rels) as f:
                    root = ET.parse(f).getroot()
                    for rel in root:
                        rid_to_file[rel.get("Id", "")] = rel.get("Target", "")
            rid = wb_sheet_rids.get(sheet_name, "")
            target_sheet_file = rid_to_file.get(rid, "").split("/")[-1] if rid else None

        # Collect tables from relevant worksheet _rels
        table_files: list[str] = []
        for rels_path in names:
            if "_rels" in rels_path and rels_path.endswith(".rels") and "worksheets" in rels_path:
                sheet_file = rels_path.split("/")[-1].replace(".rels", "")
                if target_sheet_file and sheet_file != target_sheet_file:
                    continue
                with zf.open(rels_path) as f:
                    root = ET.parse(f).getroot()
                    for rel in root:
                        target = rel.get("Target", "")
                        if "/tables/" in target:
                            table_files.append(target.split("/")[-1])

        # Read each table XML for its display name. The all-tables fallback
        # only applies when NO sheet was requested — a specific sheet with an
        # empty table_files legitimately has zero tables (previously this
        # misreported every table in the workbook for such sheets).
        list_all = not table_files and target_sheet_file is None
        for fname in names:
            if fname.startswith("xl/tables/") and fname.endswith(".xml"):
                if fname.split("/")[-1] in table_files or list_all:
                    with zf.open(fname) as f:
                        root = ET.parse(f).getroot()
                        nm = root.get("displayName") or root.get("name", "")
                        if nm:
                            tables.append(nm)
    return tables


@app.route("/api/columns", methods=["POST"])
def api_columns():
    """Return column names — header-only read, sub-100ms regardless of file size."""
    data = request.json
    path = data.get("path", "")
    sheet_name = data.get("sheet_name") or None
    table_name = data.get("table_name") or None
    header_row = _header_row_to_pandas(data.get("header_row", 1))
    try:
        if is_csv(path):
            # CSV: read only the header row
            file_bytes = read_shared_bytes(path)
            df = pd.read_csv(io.BytesIO(file_bytes), header=header_row, nrows=0, dtype=str)
            columns = list(df.columns)
        elif table_name:
            # Read column names from table XML in the ZIP — instant
            file_bytes = read_shared_bytes(path)
            columns = _table_columns_from_zip(file_bytes, table_name)
        else:
            # Read only the header row via pandas — fast even for large files
            file_bytes = read_shared_bytes(path)
            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name or 0, header=header_row, nrows=0, dtype=str)
            columns = list(df.columns)
        return jsonify({"columns": columns})
    except Exception as e:
        return jsonify({"error": friendly_read_error(e)}), 400


def _table_columns_from_zip(file_bytes: bytes, table_name: str) -> list[str]:
    """Return ordered column names for a named Excel Table from the ZIP metadata."""
    with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
        for fname in zf.namelist():
            if fname.startswith("xl/tables/") and fname.endswith(".xml"):
                with zf.open(fname) as f:
                    root = ET.parse(f).getroot()
                    nm = root.get("displayName") or root.get("name", "")
                    if nm == table_name:
                        return [c.get("name", "") for c in root.findall(".//{*}tableColumn")]
    raise ValueError(f"Table '{table_name}' not found in workbook")


@app.route("/api/workbook_map", methods=["POST"])
def api_workbook_map():
    """One-shot metadata map for the builder diagram pane: every sheet with its
    Excel Tables and header-row columns. One read_shared_bytes call, one ZIP
    pass per lookup — vs N /api/columns round trips (each re-reading the file,
    which on a locked/OneDrive workbook means a full temp copy per call)."""
    data = request.json or {}
    path = data.get("path", "")
    header_row = _header_row_to_pandas(data.get("header_row", 1))
    try:
        file_bytes = read_shared_bytes(path)
        if is_csv(path):
            df = pd.read_csv(io.BytesIO(file_bytes), header=header_row, nrows=0, dtype=str)
            return jsonify({
                "is_csv": True,
                "sheets": [{"name": None, "columns": list(df.columns), "tables": []}],
            })
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        # All sheet headers in a single pandas call (dict of empty frames).
        headers = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None,
                                header=header_row, nrows=0, dtype=str)
        sheets = []
        for name in sheet_names:
            cols = list(headers[name].columns) if name in headers else []
            tables = []
            for tname in _tables_for_sheet_from_zip(file_bytes, name):
                tables.append({"name": tname,
                               "columns": _table_columns_from_zip(file_bytes, tname)})
            sheets.append({"name": name, "columns": cols, "tables": tables})
        return jsonify({"is_csv": False, "sheets": sheets})
    except Exception as e:
        return jsonify({"error": friendly_read_error(e)}), 400


@app.route("/api/template_export")
def api_template_export():
    name = request.args.get("template", "")
    if not name:
        return jsonify({"error": "Template name required."}), 400
    try:
        t = load_template(name)
    except Exception:
        return jsonify({"error": "Template not found."}), 404
    t = dict(t)
    stype = (t.get("source", {}).get("type") or "local").strip()
    if stype == "sql":
        t["source"] = dict(t["source"])
        t["source"]["connection_id"] = ""
        t["needs_connection"] = True
    import re
    slug = re.sub(r"[^\w\-]", "_", name).lower()
    resp = app.response_class(
        json.dumps(t, indent=2),
        mimetype="application/json",
        headers={"Content-Disposition": f'attachment; filename="{slug}.waire-template.json"'},
    )
    return resp


@app.route("/api/template_import", methods=["POST"])
def api_template_import():
    data = request.get_json(silent=True)
    if not data or not isinstance(data, dict):
        return jsonify({"error": "Invalid JSON."}), 400
    errors = validate_template(data)
    if errors:
        return jsonify({"error": "; ".join(errors)}), 400
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "Template name is required."}), 400
    try:
        load_template(name)
        return jsonify({"error": f"A template named '{name}' already exists — rename it first or delete the old one."}), 400
    except Exception:
        pass
    try:
        save_template(data)
        log.log_template_save(name)
        return jsonify({"ok": True, "name": name})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/template_keys", methods=["GET"])
def api_template_keys():
    """Return key columns and view names for a named template (used by the
    link builder — views let a cross-template link target a specific view)."""
    name = request.args.get("name", "")
    if not name:
        return jsonify({"keys": [], "views": []})
    try:
        tpl = load_template(name)
        views = [v.get("name", "") for v in (tpl.get("views") or []) if v.get("name")]
        return jsonify({"keys": tpl.get("key_columns", []), "views": views})
    except Exception:
        return jsonify({"keys": [], "views": []})


@app.route("/api/save_template", methods=["POST"])
def api_save_template():
    """Save or update a template from JSON payload."""
    data = request.json
    errors = validate_template(data)
    if errors:
        return jsonify({"errors": errors}), 400
    try:
        save_template(data)
        log.log_template_save(data.get("name", ""))
        # Evict every cached source instance for this template — cache keys
        # now include sheet/table, and a single template may own many.
        src = data.get("source", {}) or {}
        stype = (src.get("type") or "local").strip()
        if stype == "sharepoint":
            prefix = f"sp::{src.get('drive_id')}::{src.get('item_id')}::"
        elif stype == "sql":
            prefix = f"sql::{src.get('connection_id', '')}::"
        else:
            prefix = f"{src.get('path', '')}::"
        for k in [k for k in _sources if k.startswith(prefix)]:
            _sources.pop(k, None)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"errors": [str(e)]}), 500


# ---------------------------------------------------------------------------
# Send-to pipeline
# ---------------------------------------------------------------------------

def _parse_send_payload():
    data = request.json or {}
    columns = data.get("columns") or []
    rows = data.get("rows") or []
    if not columns or not rows:
        raise ValueError("Nothing selected to send.")
    return (data.get("template", ""), columns, rows,
            data.get("deep_link", ""), data)


@app.route("/api/send/outlook", methods=["POST"])
def api_send_outlook():
    from core import send_format, send_outlook
    try:
        template, columns, rows, deep_link, _data = _parse_send_payload()
        html_body = send_format.build_mail_html(template, columns, rows, deep_link)
        subject = f"WAIRE LookUp — {template} ({len(rows)} results)"
        send_outlook.create_draft(subject, html_body)
        log.log_send("outlook", template, len(rows))
        return jsonify({"ok": True})
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/send/excel", methods=["POST"])
def api_send_excel():
    """Generate a fresh .xlsx of the sent rows and open it directly in Excel
    (temp file + COM) — no browser download, no manual re-opening. Never
    looks for or appends to any existing workbook."""
    from core import send_excel
    try:
        template, columns, rows, _deep_link, _data = _parse_send_payload()
        send_excel.open_in_excel(columns, rows)
        log.log_send("excel", template, len(rows))
        return jsonify({"ok": True})
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/send/teams", methods=["POST"])
def api_send_teams():
    from core import send_format, send_teams
    try:
        template, columns, rows, deep_link, data = _parse_send_payload()
        webhook_id = data.get("target") or ""
        hooks = load_settings().get("teams_webhooks", [])
        hook = next((h for h in hooks if h.get("id") == webhook_id), None)
        if not hook:
            return jsonify({"ok": False, "error": "Webhook not found."}), 400
        card = send_format.build_teams_card(template, columns, rows, deep_link)
        send_teams.post_card(hook["url"], card)
        log.log_send("teams", template, len(rows))  # never log the webhook URL
        return jsonify({"ok": True})
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/teams_webhooks", methods=["GET", "POST", "DELETE"])
def api_teams_webhooks():
    if request.method == "GET":
        hooks = load_settings().get("teams_webhooks", [])
        return jsonify([{"id": h["id"], "name": h.get("name", ""), "url_tail": h.get("url", "")[-8:]}
                        for h in hooks])
    if request.method == "DELETE":
        hid = request.args.get("id", "")
        hooks = [h for h in load_settings().get("teams_webhooks", []) if h.get("id") != hid]
        save_settings({"teams_webhooks": hooks})
        return jsonify({"ok": True})
    data = request.json or {}
    name = (data.get("name") or "").strip()
    url = (data.get("url") or "").strip()
    if not name or not url:
        return jsonify({"ok": False, "error": "Name and URL are required."}), 400
    if not url.startswith("https://"):
        return jsonify({"ok": False, "error": "Webhook URL must start with https://."}), 400
    import uuid
    hooks = load_settings().get("teams_webhooks", [])
    hid = str(uuid.uuid4())
    hooks.append({"id": hid, "name": name, "url": url})
    save_settings({"teams_webhooks": hooks})
    log.log_settings_change({"teams_webhook_added": name})  # never log the URL
    return jsonify({"ok": True, "id": hid})


# ---------------------------------------------------------------------------
# SQL Server connections + query check
# ---------------------------------------------------------------------------

@app.route("/api/sql_connections", methods=["GET", "POST", "DELETE"])
def api_sql_connections():
    from core import sql_connections
    if request.method == "GET":
        return jsonify({"connections": sql_connections.list_connections()})
    if request.method == "DELETE":
        cid = request.args.get("id", "")
        sql_connections.delete_connection(cid)
        return jsonify({"ok": True})
    # POST — create or update
    data = request.json or {}
    try:
        cid = sql_connections.save_connection(
            connection_id=data.get("id") or None,
            name=(data.get("name") or "").strip(),
            server=(data.get("server") or "").strip(),
            port=int(data.get("port") or 1433),
            database=(data.get("database") or "").strip(),
            username=(data.get("username") or "").strip(),
            password=data.get("password") or "",
        )
        return jsonify({"ok": True, "id": cid})
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/sql_check", methods=["POST"])
def api_sql_check():
    """Run the query wrapped as TOP 0; return columns or the raw error."""
    from connectors.sql_query import check_query
    from core import sql_connections, sql_credentials
    data = request.json or {}
    conn_id = (data.get("connection_id") or "").strip()
    query = (data.get("query") or "").strip()
    if not conn_id:
        return jsonify({"ok": False, "error": "No SQL connection selected."}), 400
    if not query:
        return jsonify({"ok": False, "error": "Query is empty."}), 400
    conn = sql_connections.load_connection(conn_id)
    if not conn:
        return jsonify({"ok": False, "error": "Connection not found."}), 404
    cred = sql_credentials.load_credential(conn.get("credential_id", ""))
    if not cred:
        return jsonify({"ok": False, "error": "No stored password for this connection."}), 400
    username, password = cred
    full_conn = {**conn, "username": username, "password": password}
    result = check_query(full_conn, query)
    return jsonify(result)


# ---------------------------------------------------------------------------
# Template builder — page routes
# ---------------------------------------------------------------------------

@app.route("/restart", methods=["POST"])
def do_restart():
    def _restart():
        time.sleep(0.5)  # let response flush
        # Detached helper: waits for this process to exit, then relaunches
        helper = (
            "import time, subprocess; "
            f"time.sleep(2); "
            f"subprocess.Popen({[sys.executable] + sys.argv!r})"
        )
        flags = 0x08000000 if sys.platform == "win32" else 0  # CREATE_NO_WINDOW
        subprocess.Popen([sys.executable, "-c", helper], creationflags=flags)
        os._exit(0)
    threading.Thread(target=_restart, daemon=False).start()
    return """<!doctype html>
<html><head>
<meta charset="UTF-8">
<title>Restarting…</title>
<style>body{font-family:system-ui,sans-serif;display:flex;align-items:center;justify-content:center;height:100vh;margin:0;background:#f5f5f5;}
.box{text-align:center;color:#333;}.spinner{width:32px;height:32px;border:3px solid #ddd;border-top-color:#1a3a5c;border-radius:50%;animation:spin .8s linear infinite;margin:12px auto;}
@keyframes spin{to{transform:rotate(360deg);}}</style>
<script>
setTimeout(function poll(){
  fetch('/').then(r=>{ if(r.ok) location.href='/'; else setTimeout(poll,400); }).catch(()=>setTimeout(poll,400));
}, 1200);
</script>
</head><body><div class="box"><div class="spinner"></div><p>Restarting server…</p></div></body></html>"""


@app.route("/templates/new")
def template_new():
    return render_template("template_builder.html", template=None, templates=list_templates())


@app.route("/templates/<name>/edit")
def template_edit(name):
    try:
        t = load_template(name)
    except Exception:
        t = None
    return render_template("template_builder.html", template=t, templates=list_templates())


@app.route("/templates/<name>/delete", methods=["POST"])
def template_delete(name):
    delete_template(name)
    return redirect(url_for("index"))


def _preload_sources() -> None:
    """Warm the parquet cache for all file-based templates."""
    try:
        templates = templates_store.list_templates()
    except Exception:
        return
    for name in templates:
        try:
            tpl = templates_store.load_template(name)
            if not tpl:
                continue
            stype = (tpl.get("source", {}).get("type") or "local").strip()
            if stype == "sql":
                continue
            groups = group_views_by_source(tpl)
            for g in groups:
                src = _get_view_group_source(tpl, g)
                if src.is_stale():
                    src.load()
        except Exception:
            pass


def start_background() -> None:
    """Start the source poller + do a snapshot cleanup. Idempotent."""
    try:
        snapshot_store.cleanup_snapshots(config.SNAPSHOT_TTL_HOURS)
    except Exception:
        pass
    start_poller()
    threading.Thread(target=_preload_sources, daemon=True).start()


def ensure_single_instance(port: int, host: str = "127.0.0.1") -> None:
    """Kill whatever's already listening on our port before binding.

    Dev workflow tends to accumulate stray `python app.py` processes across
    restarts; the older one can keep serving stale code after edits even
    though a newer one was started. This guarantees a fresh start always
    wins the port, so there is only ever one server and Restart is reliable.
    """
    if sys.platform != "win32":
        return

    def _owner_pid():
        try:
            out = subprocess.run(
                ["netstat", "-ano"], capture_output=True, text=True, timeout=5
            ).stdout
        except Exception:
            return None
        for line in out.splitlines():
            parts = line.split()
            if len(parts) >= 5 and parts[0] == "TCP" and parts[1].endswith(f":{port}") and parts[3] == "LISTENING":
                try:
                    return int(parts[-1])
                except ValueError:
                    return None
        return None

    import socket

    for _ in range(6):
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        try:
            s.bind((host, port))
            s.close()
            return
        except OSError:
            s.close()
        pid = _owner_pid()
        if pid and pid != os.getpid():
            subprocess.run(["taskkill", "/PID", str(pid), "/F"], capture_output=True, timeout=5)
        time.sleep(0.5)


if __name__ == "__main__":
    ensure_single_instance(2305)
    start_background()
    app.run(host="127.0.0.1", port=2305, debug=False)
