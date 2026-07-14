import io
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter, range_boundaries

from core.fileio import is_csv, read_shared_bytes

from .base import DataSource


class SyncedFileSource(DataSource):
    def __init__(
        self,
        path: str,
        sheet_name: str | None = None,
        table_name: str | None = None,
        header_row: int = 0,
    ):
        self._path = Path(path)
        self._sheet_name = sheet_name
        self._table_name = table_name
        self._header_row = header_row

        self._df: pd.DataFrame | None = None
        self._cached_mtime: float | None = None

    def _current_mtime(self) -> float:
        return self._path.stat().st_mtime

    def load(self) -> pd.DataFrame:
        mtime = self._current_mtime()
        data = read_shared_bytes(self._path)

        if is_csv(self._path):
            self._df = pd.read_csv(
                io.BytesIO(data),
                header=self._header_row,
                dtype=str,
            )
        elif self._table_name:
            self._df = self._load_table(data)
        elif self._sheet_name is not None:
            self._df = pd.read_excel(
                io.BytesIO(data),
                sheet_name=self._sheet_name,
                header=self._header_row,
                dtype=str,
            )
        else:
            self._df = pd.read_excel(
                io.BytesIO(data),
                sheet_name=0,
                header=self._header_row,
                dtype=str,
            )

        self._df = self._df.fillna("")
        self._cached_mtime = mtime
        return self._df

    def _load_table(self, data: bytes) -> pd.DataFrame:
        """Load a named Excel Table via ZIP metadata + pd.read_excel (avoids slow openpyxl cell iteration)."""
        ref, sheet_name = self._find_table_ref(data, self._table_name)
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        usecols = f"{get_column_letter(min_col)}:{get_column_letter(max_col)}"
        skiprows = min_row - 1          # rows before the header row
        nrows = max_row - min_row        # data rows (excluding header)
        use_sheet = sheet_name if sheet_name else (self._sheet_name or 0)
        df = pd.read_excel(
            io.BytesIO(data),
            sheet_name=use_sheet,
            skiprows=skiprows,
            nrows=nrows,
            usecols=usecols,
            header=0,
            dtype=str,
        )
        return df.fillna("")

    @staticmethod
    def _find_table_ref(data: bytes, table_name: str) -> tuple[str, str | None]:
        """Return (cell_ref, sheet_name) for a named table by reading the xlsx ZIP directly."""
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            names = zf.namelist()

            # Build a map: table file basename → sheet name via worksheet _rels
            table_file_to_sheet: dict[str, str] = {}
            wb_sheet_map: dict[str, str] = {}  # sheet rId → sheet name

            # Parse workbook.xml to get sheet names
            if "xl/workbook.xml" in names:
                with zf.open("xl/workbook.xml") as f:
                    root = ET.parse(f).getroot()
                    for sheet in root.findall(".//{*}sheet"):
                        rid = sheet.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", "")
                        sname = sheet.get("name", "")
                        wb_sheet_map[rid] = sname

            # Parse workbook.xml.rels to map rId → sheet file
            rid_to_file: dict[str, str] = {}
            wb_rels = "xl/_rels/workbook.xml.rels"
            if wb_rels in names:
                with zf.open(wb_rels) as f:
                    root = ET.parse(f).getroot()
                    for rel in root:
                        rid_to_file[rel.get("Id", "")] = rel.get("Target", "")

            # For each worksheet, check its _rels for table references
            for rid, sname in wb_sheet_map.items():
                sheet_file = rid_to_file.get(rid, "")
                if not sheet_file:
                    continue
                sheet_basename = sheet_file.split("/")[-1]
                rels_path = f"xl/worksheets/_rels/{sheet_basename}.rels"
                if rels_path in names:
                    with zf.open(rels_path) as f:
                        root = ET.parse(f).getroot()
                        for rel in root:
                            target = rel.get("Target", "")
                            if "/tables/" in target:
                                tbl_file = target.split("/")[-1]
                                table_file_to_sheet[tbl_file] = sname

            # Now find the table XML matching our name
            for fname in names:
                if fname.startswith("xl/tables/") and fname.endswith(".xml"):
                    with zf.open(fname) as f:
                        root = ET.parse(f).getroot()
                        nm = root.get("displayName") or root.get("name", "")
                        if nm == table_name:
                            ref = root.get("ref", "")
                            tbl_basename = fname.split("/")[-1]
                            sheet = table_file_to_sheet.get(tbl_basename)
                            return ref, sheet

        raise ValueError(f"Table '{table_name}' not found in workbook")

    def columns(self) -> list[str]:
        if self._df is None:
            self.load()
        return list(self._df.columns)

    def source_timestamp(self) -> datetime:
        if self._cached_mtime is None:
            return datetime.fromtimestamp(self._current_mtime())
        return datetime.fromtimestamp(self._cached_mtime)

    def is_stale(self) -> bool:
        if self._cached_mtime is None:
            return True
        try:
            return self._current_mtime() != self._cached_mtime
        except OSError:
            return True

    @property
    def dataframe(self) -> pd.DataFrame | None:
        return self._df
