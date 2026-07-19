"""/api/workbook_map — one-shot sheet/table/column metadata for the builder diagram."""

import json
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.worksheet.table import Table
import pandas as pd
import pytest


@pytest.fixture
def client():
    from app import app
    app.config["TESTING"] = True
    return app.test_client()


@pytest.fixture
def two_sheet_xlsx(tmp_path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "People"
    ws1.append(["ID", "Name", "City"])
    ws1.append(["1", "Alice", "LA"])
    ws2 = wb.create_sheet("Orders")
    ws2.append(["OrderID", "Amount"])
    ws2.append(["A1", "100"])
    ws2.append(["A2", "200"])
    ws2.add_table(Table(displayName="TOrders", ref="A1:B3"))
    path = tmp_path / "book.xlsx"
    wb.save(path)
    return str(path)


def _post(client, path, header_row=1):
    return client.post("/api/workbook_map", json={"path": path, "header_row": header_row})


def test_workbook_map_two_sheets(client, two_sheet_xlsx):
    rv = _post(client, two_sheet_xlsx)
    assert rv.status_code == 200
    j = json.loads(rv.data)
    assert j["is_csv"] is False
    assert [s["name"] for s in j["sheets"]] == ["People", "Orders"]
    assert j["sheets"][0]["columns"] == ["ID", "Name", "City"]
    assert j["sheets"][1]["columns"] == ["OrderID", "Amount"]


def test_workbook_map_table_columns(client, two_sheet_xlsx):
    j = json.loads(_post(client, two_sheet_xlsx).data)
    orders = next(s for s in j["sheets"] if s["name"] == "Orders")
    assert orders["tables"] == [{"name": "TOrders", "columns": ["OrderID", "Amount"]}]
    people = next(s for s in j["sheets"] if s["name"] == "People")
    assert people["tables"] == []


def test_workbook_map_csv(client, tmp_path):
    csv = tmp_path / "data.csv"
    pd.DataFrame({"A": ["1"], "B": ["2"], "C": ["3"]}).to_csv(csv, index=False)
    rv = _post(client, str(csv))
    assert rv.status_code == 200
    j = json.loads(rv.data)
    assert j["is_csv"] is True
    assert len(j["sheets"]) == 1
    assert j["sheets"][0]["name"] is None
    assert j["sheets"][0]["columns"] == ["A", "B", "C"]
    assert j["sheets"][0]["tables"] == []


def test_workbook_map_header_row(client, tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["junk", "junk2"])
    ws.append(["Real1", "Real2"])
    path = tmp_path / "hdr.xlsx"
    wb.save(path)
    j = json.loads(_post(client, str(path), header_row=2).data)
    assert j["sheets"][0]["columns"] == ["Real1", "Real2"]


def test_workbook_map_missing_file(client, tmp_path):
    rv = _post(client, str(tmp_path / "nope.xlsx"))
    assert rv.status_code == 400
    assert "error" in json.loads(rv.data)
