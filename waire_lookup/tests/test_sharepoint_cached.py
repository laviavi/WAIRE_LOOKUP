import io
import sys
import zipfile
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import pytest

import config
from connectors.sharepoint_cached import SharePointCachedSource
from core import graph_client, source_sync


def _valid_xlsx_bytes():
    """A minimal xlsx with one visible sheet containing 'a' and '1'."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "a"
    ws["A2"] = "1"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def env(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "SOURCE_STATUS_FILE", tmp_path / "status.json")
    monkeypatch.setattr(config, "SOURCE_CACHE_DIR", tmp_path / "cache")
    monkeypatch.setattr(config, "LOG_FILE", tmp_path / "log.txt")
    (tmp_path / "cache").mkdir()
    from core import graph_auth
    monkeypatch.setattr(graph_auth, "is_configured", lambda: True)
    monkeypatch.setattr(graph_auth, "get_token_silent", lambda: "T")
    body = _valid_xlsx_bytes()

    def _md(drive_id, item_id, token, session=None):
        return graph_client.DriveItemRef(drive_id, item_id, "f.xlsx", "e1", "2026-07-10T00:00:00Z")

    def _dl(drive_id, item_id, token, dest_path, session=None):
        dest_path.parent.mkdir(parents=True, exist_ok=True)
        dest_path.write_bytes(body)

    monkeypatch.setattr(graph_client, "get_item_metadata", _md)
    monkeypatch.setattr(graph_client, "download_item", _dl)
    return tmp_path


def test_cache_path_derived(env):
    p = source_sync.cache_path_for("01ABC-item", "book.xlsx")
    assert p.suffix == ".xlsx"
    assert "01ABC-item" in p.name


def test_ensure_cached_triggers_download(env):
    src = SharePointCachedSource(
        drive_id="drv", item_id="itm", name="f.xlsx",
        sheet_name="Sheet1", header_row=0, template_name="t1",
    )
    assert not src._path.exists()
    src.ensure_cached()
    assert src._path.exists()


def test_load_reads_from_cache(env):
    src = SharePointCachedSource(
        drive_id="drv", item_id="itm2", name="f.xlsx",
        sheet_name="Sheet1", header_row=0, template_name="t2",
    )
    df = src.load()
    assert list(df.columns) == ["a"]
    assert df["a"].tolist() == ["1"]
