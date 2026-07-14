import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import io

import openpyxl

from core.send_excel import build_workbook


def test_build_workbook_header_and_rows():
    data = build_workbook(["A", "B"], [["1", "2"], ["3", "4"]])
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["A", "B"]
    assert [c.value for c in ws[2]] == ["1", "2"]
    assert [c.value for c in ws[3]] == ["3", "4"]


def test_build_workbook_header_only_when_no_rows():
    data = build_workbook(["A", "B"], [])
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    assert ws.max_row == 1
    assert [c.value for c in ws[1]] == ["A", "B"]
