import json
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
import pytest


@pytest.fixture
def clean_env(tmp_path, monkeypatch):
    import config
    monkeypatch.setattr(config, "LOOKUP_TEMPLATES_DIR", tmp_path / "templates")
    monkeypatch.setattr(config, "EXPORTS_DIR", tmp_path / "exports")
    monkeypatch.setattr(config, "LOGS_DIR", tmp_path / "logs")
    monkeypatch.setattr(config, "LOG_FILE", tmp_path / "logs" / "lookups.log")
    monkeypatch.setattr(config, "SETTINGS_FILE", tmp_path / "settings.json")
    monkeypatch.setattr(config, "SNAPSHOTS_DIR", tmp_path / "snapshots")
    monkeypatch.setattr(config, "SOURCE_STATUS_FILE", tmp_path / "source_status.json")
    for d in ("templates", "exports", "logs", "snapshots"):
        (tmp_path / d).mkdir()
    return tmp_path


def _write_template(dir_: Path, name: str, csv_path: Path):
    (dir_ / f"{name}.json").write_text(json.dumps({
        "name": name,
        "source": {"path": str(csv_path), "sheet_name": None, "table_name": None, "header_row": 1},
        "key_columns": ["City"],
        "result_columns": ["PropertyID", "Address", "City"],
        "labels": {},
        "default_filter": None,
        "default_match_mode": "partial",
        "schema_version": 1,
    }), encoding="utf-8")


def test_send_outlook_empty_payload_400(clean_env):
    import app as _app
    client = _app.app.test_client()
    rv = client.post("/api/send/outlook", json={})
    assert rv.status_code == 400
    assert "Nothing selected" in rv.get_json()["error"]


def test_send_outlook_happy_path(clean_env, monkeypatch):
    import app as _app
    from core import send_outlook
    captured = {}

    def fake_create_draft(subject, html_body):
        captured["subject"] = subject
        captured["html_body"] = html_body

    monkeypatch.setattr(send_outlook, "create_draft", fake_create_draft)

    client = _app.app.test_client()
    rv = client.post("/api/send/outlook", json={
        "template": "costar",
        "columns": ["PropertyID", "City"],
        "rows": [["0012345", "Corona"]],
        "deep_link": "",
    })
    assert rv.status_code == 200
    assert rv.get_json()["ok"] is True
    assert "0012345" in captured["html_body"]


def test_send_teams_unknown_webhook_400(clean_env):
    import app as _app
    client = _app.app.test_client()
    rv = client.post("/api/send/teams", json={
        "template": "costar",
        "columns": ["A"],
        "rows": [["1"]],
        "target": "nonexistent-id",
    })
    assert rv.status_code == 400


def test_send_excel_empty_payload_400(clean_env):
    import app as _app
    client = _app.app.test_client()
    rv = client.post("/api/send/excel", json={})
    assert rv.status_code == 400
    assert "Nothing selected" in rv.get_json()["error"]


def test_send_excel_downloads_generated_workbook(clean_env):
    import io
    import openpyxl
    import app as _app
    client = _app.app.test_client()
    rv = client.post("/api/send/excel", json={
        "template": "costar",
        "columns": ["PropertyID", "City"],
        "rows": [["0012345", "Corona"]],
    })
    assert rv.status_code == 200
    assert "attachment" in rv.headers.get("Content-Disposition", "")
    wb = openpyxl.load_workbook(io.BytesIO(rv.data))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["PropertyID", "City"]
    assert [c.value for c in ws[2]] == ["0012345", "Corona"]


def test_deep_link_prefill_and_auto_run(clean_env):
    import app as _app
    csv = clean_env / "ac.csv"
    pd.DataFrame({"City": ["LA"]}).to_csv(csv, index=False)
    _write_template(clean_env / "templates", "ac", csv)

    client = _app.app.test_client()

    rv = client.get("/?template=ac&key_0=abc&mode=partial")
    assert rv.status_code == 200
    html = rv.data.decode("utf-8")
    assert ">abc</textarea>" in html
    assert 'selected>Partial' in html
    assert "f.submit()" not in html  # no run=1 -> no auto-run snippet

    rv2 = client.get("/?template=ac&key_0=abc&mode=partial&run=1")
    html2 = rv2.data.decode("utf-8")
    assert "f.submit()" in html2
